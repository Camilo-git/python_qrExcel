import os
import re
import sys
from typing import List, Optional, Tuple

try:
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox
except Exception as e:
    print(
        "Error importando Tkinter. En Linux puede que necesites instalar python3-tk (o tk).",
        file=sys.stderr,
    )
    raise

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import qrcode


class ExcelPreviewApp(tk.Tk):
    """App para cargar un .xlsx, mostrar 2 filas y generar QR (col B) con título (col A)."""

    def __init__(self) -> None:
        super().__init__()
        self.title("Vista de 2 filas de Excel + QR")
        self.geometry("900x500")
        self.minsize(700, 350)

        # Estado
        self.tree: Optional[ttk.Treeview] = None
        self.status_var: tk.StringVar = tk.StringVar(
            value="Selecciona un archivo .xlsx para ver 2 filas y generar QR."
        )
        self.current_excel_path: Optional[str] = None
        self.skip_header_var: tk.BooleanVar = tk.BooleanVar(value=True)

        self._build_ui()

    def _build_ui(self) -> None:
        # Barra superior
        top = ttk.Frame(self)
        top.pack(side=tk.TOP, fill=tk.X, padx=10, pady=10)

        self.load_btn = ttk.Button(top, text="Cargar Excel (.xlsx)", command=self.on_load_excel)
        self.load_btn.pack(side=tk.LEFT)

        self.clear_btn = ttk.Button(top, text="Limpiar", command=self.clear_table)
        self.clear_btn.pack(side=tk.LEFT, padx=(10, 0))

        skip_cb = ttk.Checkbutton(
            top,
            text="Omitir primera fila (encabezado)",
            variable=self.skip_header_var,
        )
        skip_cb.pack(side=tk.LEFT, padx=(10, 0))

        self.qr_btn = ttk.Button(top, text="Generar QR", command=self.on_generate_qr)
        self.qr_btn.pack(side=tk.LEFT, padx=(10, 0))

        self.quit_btn = ttk.Button(top, text="Salir", command=self.destroy)
        self.quit_btn.pack(side=tk.RIGHT)

        # Tabla con scrollbars
        table_frame = ttk.Frame(self)
        table_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        tree = ttk.Treeview(table_frame, columns=(), show="headings")
        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        self.tree = tree

        # Barra de estado
        status = ttk.Label(self, textvariable=self.status_var, anchor="w")
        status.pack(side=tk.BOTTOM, fill=tk.X, padx=10, pady=(0, 8))

    def clear_table(self) -> None:
        if not self.tree:
            return
        for col in self.tree["columns"]:
            self.tree.heading(col, text="")
            self.tree.column(col, width=0, stretch=False)
        self.tree["columns"] = ()
        for item in self.tree.get_children():
            self.tree.delete(item)
        self.status_var.set("Tabla vacía. Carga un archivo .xlsx.")

    def on_load_excel(self) -> None:
        path = filedialog.askopenfilename(
            title="Seleccionar archivo de Excel",
            filetypes=[
                ("Excel Workbook", "*.xlsx"),
                ("Excel Macro-Enabled", "*.xlsm"),
                ("Todos los archivos", "*.*"),
            ],
        )
        if not path:
            return
        try:
            rows = self.read_two_columns_all_rows(path)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer el archivo:\n{e}")
            return
        if not rows:
            messagebox.showwarning("Sin datos", "No se encontraron filas en la hoja activa.")
            self.clear_table()
            return
        self.populate_table(rows)
        self.status_var.set(f"Archivo cargado: {path} — Mostrando {len(rows)} fila(s).")
        self.current_excel_path = path

    def read_first_two_rows(self, path: str) -> List[List[Optional[object]]]:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            raise ValueError("El libro de Excel no tiene una hoja activa.")
        out: List[List[Optional[object]]] = []
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            out.append(list(row) if row is not None else [])
            if idx >= 1:
                break
        wb.close()
        return out

    def read_two_columns_all_rows(self, path: str) -> List[List[Optional[object]]]:
        """Lee todas las filas de la hoja activa, únicamente columnas A y B."""
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            raise ValueError("El libro de Excel no tiene una hoja activa.")
        out: List[List[Optional[object]]] = []
        for row in ws.iter_rows(values_only=True):
            if row is None:
                out.append([None, None])
                continue
            a = row[0] if len(row) > 0 else None
            b = row[1] if len(row) > 1 else None
            out.append([a, b])
        wb.close()
        return out

    def populate_table(self, rows: List[List[Optional[object]]]) -> None:
        if not self.tree:
            return
        self.clear_table()
        # Mostrar solo 2 columnas (A y B) y todas las filas
        max_cols = 2 if rows else 0
        if max_cols == 0:
            messagebox.showwarning("Sin datos", "Las primeras filas están vacías.")
            return
        columns = [get_column_letter(i + 1) for i in range(max_cols)]
        self.tree["columns"] = columns
        for col in columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=140, anchor="center")
        for r in rows:
            normalized = ["" if v is None else str(v) for v in (r + [None] * (max_cols - len(r)))]
            self.tree.insert("", tk.END, values=normalized)

    def on_generate_qr(self) -> None:
        if not self.current_excel_path:
            messagebox.showwarning("Sin archivo", "Primero carga un archivo Excel para generar los QR.")
            return
        try:
            count, skipped = self.generate_qr_images(self.current_excel_path, skip_header=self.skip_header_var.get())
        except Exception as e:
            messagebox.showerror("Error", f"No se pudieron generar los QR:\n{e}")
            return
        msg = f"Se generaron {count} QR"
        if skipped:
            msg += f" — {skipped} fila(s) omitidas por datos incompletos."
        messagebox.showinfo("QR listos", msg)
        self.status_var.set(msg)

    def _safe_filename(self, name: str) -> str:
        name = name.strip()
        name = re.sub(r"[^\w\-\. ]+", "_", name)
        name = re.sub(r"\s+", "_", name)
        return name or "qr"

    def generate_qr_images(self, path: str, skip_header: bool = False) -> Tuple[int, int]:
        # Guardar los QR junto al ejecutable si está congelado (PyInstaller),
        # de lo contrario junto al script.
        if getattr(sys, "frozen", False):  # ejecutable PyInstaller
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
        out_dir = os.path.join(base_dir, "img")
        os.makedirs(out_dir, exist_ok=True)

        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            wb.close()
            raise ValueError("El libro de Excel no tiene una hoja activa.")

        generated = 0
        skipped = 0
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if skip_header and r_idx == 1:
                continue
            title = row[0] if row and len(row) > 0 else None
            url = row[1] if row and len(row) > 1 else None
            if title is None or url is None:
                skipped += 1
                continue
            title_str = str(title).strip()
            url_str = str(url).strip()
            if not title_str or not url_str:
                skipped += 1
                continue

            fname_base = self._safe_filename(title_str)
            dest = os.path.join(out_dir, f"{fname_base}.png")
            suffix = 1
            while os.path.exists(dest):
                dest = os.path.join(out_dir, f"{fname_base}_{suffix}.png")
                suffix += 1

            qr = qrcode.QRCode(version=1, box_size=10, border=4)
            qr.add_data(url_str)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            with open(dest, "wb") as f:
                img.save(f)
            generated += 1

        wb.close()
        return generated, skipped


if __name__ == "__main__":
    app = ExcelPreviewApp()
    app.mainloop()
